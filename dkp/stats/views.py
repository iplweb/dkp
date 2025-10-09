from datetime import datetime, timedelta
from django.shortcuts import render
from django.db.models import Q
from django.utils import timezone
from django.contrib.sites.shortcuts import get_current_site
from django.http import HttpResponse
from django.utils.translation import gettext as _, get_language
from django.utils.safestring import mark_safe
import json
from comms.models import MessageLog
from hospital.models import OperatingRoom, Ward
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_translated_name(obj):
    """
    Get translated name for a model object with fallback support.

    Tries to get the name in the current language, falls back to name_en,
    then to the base name field.

    Args:
        obj: Model instance with name field (Ward, OperatingRoom, etc.)

    Returns:
        str: The best available name for the object
    """
    # Get current language
    current_lang = get_language()

    # Try current language's translated field
    if current_lang:
        lang_code = current_lang.split('-')[0]  # Convert 'en-us' to 'en'
        translated_field = f'name_{lang_code}'
        if hasattr(obj, translated_field):
            translated_name = getattr(obj, translated_field, None)
            if translated_name:
                return translated_name

    # Fallback to English translation
    if hasattr(obj, 'name_en') and obj.name_en:
        return obj.name_en

    # Final fallback to base name field
    return obj.name if hasattr(obj, 'name') else str(obj)


def get_date_range(period, date_str=None):
    """
    Calculate start and end dates for the given period.

    Args:
        period: 'day', 'week', 'month', or 'quarter'
        date_str: Date string in YYYY-MM-DD format (optional, defaults to today)

    Returns:
        Tuple of (start_date, end_date, display_name)
    """
    if date_str:
        try:
            reference_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            reference_date = timezone.now().date()
    else:
        reference_date = timezone.now().date()

    if period == 'day':
        start_date = reference_date
        end_date = reference_date + timedelta(days=1)
        display_name = reference_date.strftime('%Y-%m-%d')

    elif period == 'week':
        # Start from Monday
        start_date = reference_date - timedelta(days=reference_date.weekday())
        end_date = start_date + timedelta(days=7)
        display_name = f"Week {start_date.strftime('%Y-%m-%d')} to {(end_date - timedelta(days=1)).strftime('%Y-%m-%d')}"

    elif period == 'month':
        start_date = reference_date.replace(day=1)
        # Next month
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1)
        display_name = start_date.strftime('%B %Y')

    elif period == 'quarter':
        # Calculate quarter (Q1: Jan-Mar, Q2: Apr-Jun, Q3: Jul-Sep, Q4: Oct-Dec)
        quarter = (reference_date.month - 1) // 3 + 1
        quarter_start_month = (quarter - 1) * 3 + 1
        start_date = reference_date.replace(month=quarter_start_month, day=1)

        # End of quarter
        if quarter == 4:
            end_date = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date = start_date.replace(month=quarter_start_month + 3)

        display_name = f"Q{quarter} {start_date.year}"

    else:
        # Default to day
        start_date = reference_date
        end_date = reference_date + timedelta(days=1)
        display_name = reference_date.strftime('%Y-%m-%d')

    # Convert to datetime for filtering
    start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.min.time()))

    return start_datetime, end_datetime, display_name


def match_or_busy_periods(patient_requested_messages, patient_in_or_messages, surgery_done_messages):
    """
    Match CAN_ACCEPT_PATIENTS -> PATIENT_IN_THE_OR -> SURGERY_DONE to calculate OR busy time.

    Args:
        patient_requested_messages: QuerySet of CAN_ACCEPT_PATIENTS messages
        patient_in_or_messages: QuerySet of PATIENT_IN_THE_OR messages
        surgery_done_messages: QuerySet of SURGERY_DONE messages

    Returns:
        List of dictionaries with matched busy periods and incomplete entries
    """
    journeys = []

    # Group messages by operating room for time between calls calculation
    prev_surgery_done_by_or = {}

    for patient_requested in patient_requested_messages:
        # Find matching PATIENT_IN_THE_OR (optional middle step)
        # Same operating room, sent after CAN_ACCEPT_PATIENTS, within 24 hours
        matching_patient_in_or = patient_in_or_messages.filter(
            operating_room_id=patient_requested.operating_room_id,
            sent_at__gt=patient_requested.sent_at,
            sent_at__lt=patient_requested.sent_at + timedelta(hours=24)
        ).order_by('sent_at').first()

        # Find matching SURGERY_DONE
        # Same operating room, sent after CAN_ACCEPT_PATIENTS, within 24 hours
        matching_done = surgery_done_messages.filter(
            operating_room_id=patient_requested.operating_room_id,
            sent_at__gt=patient_requested.sent_at,
            sent_at__lt=patient_requested.sent_at + timedelta(hours=24)
        ).order_by('sent_at').first()

        # Calculate time between calls (from previous surgery done to this request)
        # Find the most recent SURGERY_DONE before this CAN_ACCEPT_PATIENTS
        previous_done = surgery_done_messages.filter(
            operating_room_id=patient_requested.operating_room_id,
            sent_at__lt=patient_requested.sent_at
        ).order_by('-sent_at').first()

        # Try to get the operating room, handle if it doesn't exist
        try:
            operating_room = OperatingRoom.objects.get(id=patient_requested.operating_room_id)
        except OperatingRoom.DoesNotExist:
            # Create a placeholder for deleted/non-existent OR
            class DeletedOR:
                def __init__(self, operating_room_id):
                    self.id = operating_room_id
                    self.name = f"Unexistent OR (ID: {operating_room_id})"
            operating_room = DeletedOR(patient_requested.operating_room_id)

        journey = {
            'patient_requested': patient_requested,
            'patient_in_or': matching_patient_in_or,
            'surgery_done': matching_done,
            'incomplete': matching_done is None,
            'operating_room': operating_room,
        }

        # Calculate time between calls
        if previous_done:
            time_between = patient_requested.sent_at - previous_done.sent_at
            journey['time_between_calls'] = time_between
            journey['time_between_calls_minutes'] = int(time_between.total_seconds() / 60)
        else:
            journey['time_between_calls'] = None
            journey['time_between_calls_minutes'] = None

        # Calculate wait time (request to patient in OR)
        if matching_patient_in_or:
            wait_time = matching_patient_in_or.sent_at - patient_requested.sent_at
            journey['wait_time'] = wait_time
            journey['wait_time_minutes'] = int(wait_time.total_seconds() / 60)
        else:
            journey['wait_time'] = None
            journey['wait_time_minutes'] = None

        # Calculate operation time (patient in OR to surgery done)
        if matching_patient_in_or and matching_done:
            operation_time = matching_done.sent_at - matching_patient_in_or.sent_at
            journey['operation_time'] = operation_time
            journey['operation_time_minutes'] = int(operation_time.total_seconds() / 60)
        else:
            journey['operation_time'] = None
            journey['operation_time_minutes'] = None

        # Calculate total OR time (request to surgery done)
        if matching_done:
            total_time = matching_done.sent_at - patient_requested.sent_at
            journey['total_time'] = total_time
            journey['total_time_minutes'] = int(total_time.total_seconds() / 60)
            # Keep 'duration' for backwards compatibility
            journey['duration'] = total_time
            journey['duration_minutes'] = int(total_time.total_seconds() / 60)
        else:
            journey['total_time'] = None
            journey['total_time_minutes'] = None
            journey['duration'] = None
            journey['duration_minutes'] = None

        journeys.append(journey)

    return journeys


def calculate_statistics(journeys):
    """
    Calculate summary statistics from journeys.

    Args:
        journeys: List of journey dictionaries

    Returns:
        Dictionary with summary statistics
    """
    total_journeys = len(journeys)
    completed_journeys = [j for j in journeys if not j['incomplete']]
    incomplete_journeys = [j for j in journeys if j['incomplete']]

    total_completed = len(completed_journeys)
    total_incomplete = len(incomplete_journeys)

    if total_completed > 0:
        total_minutes = sum(j['duration_minutes'] for j in completed_journeys)
        avg_minutes = total_minutes / total_completed
        total_hours = total_minutes / 60
        avg_hours = avg_minutes / 60
    else:
        total_minutes = 0
        avg_minutes = 0
        total_hours = 0
        avg_hours = 0

    return {
        'total_journeys': total_journeys,
        'total_completed': total_completed,
        'total_incomplete': total_incomplete,
        'total_minutes': int(total_minutes),
        'avg_minutes': int(avg_minutes),
        'total_hours': round(total_hours, 2),
        'avg_hours': round(avg_hours, 2),
    }


def dashboard(request, period='day', date_str=None):
    """
    Main statistics dashboard view.

    Shows statistics for the current hospital based on Site.
    Supports filtering by operating rooms and wards via URL parameters.
    """
    # Get current hospital from Site
    current_site = get_current_site(request)
    try:
        hospital = current_site.hospital
    except:
        # No hospital configured for this site
        return render(request, 'stats/dashboard.html', {
            'error': 'No hospital configured for this site.'
        })

    # Get date range
    start_date, end_date, display_name = get_date_range(period, date_str)

    # Get current language for sorting
    current_lang = get_language()

    # Determine which name field to use for sorting based on language
    if current_lang and current_lang.startswith('pl'):
        name_field = 'name_pl'
    else:
        name_field = 'name_en'

    # Get all operating rooms and wards for this hospital
    # Sort by 'sort' field first (nulls last), then by translated name
    from django.db.models import F
    all_ors = list(OperatingRoom.objects.filter(hospital=hospital).order_by(
        F('sort').asc(nulls_last=True), name_field, 'name'
    ))
    all_wards = list(Ward.objects.filter(hospital=hospital).order_by(
        F('sort').asc(nulls_last=True), name_field, 'name'
    ))

    # Parse filter parameters from URL
    or_ids_param = request.GET.get('or_ids', '')
    ward_ids_param = request.GET.get('ward_ids', '')

    # Convert to lists of integers, handling empty strings and invalid values
    selected_or_ids = []
    if or_ids_param:
        try:
            selected_or_ids = [int(x) for x in or_ids_param.split(',') if x.strip().isdigit()]
        except (ValueError, AttributeError):
            selected_or_ids = []

    selected_ward_ids = []
    if ward_ids_param:
        try:
            selected_ward_ids = [int(x) for x in ward_ids_param.split(',') if x.strip().isdigit()]
        except (ValueError, AttributeError):
            selected_ward_ids = []

    # Default: if no filters specified, select all
    if not selected_or_ids:
        selected_or_ids = [or_obj.id for or_obj in all_ors]
    if not selected_ward_ids:
        selected_ward_ids = [ward.id for ward in all_wards]

    # Build filter query for messages
    # Logic: (OR1 OR OR2 OR ...) AND (Ward1 OR Ward2 OR ...)
    # This means show messages that involve ANY of the selected ORs AND ANY of the selected wards

    messages = MessageLog.objects.filter(
        hospital=hospital,
        sent_at__gte=start_date,
        sent_at__lt=end_date
    ).select_related('sender_role', 'recipient_role')

    # Check if filters are actually active (not all selected)
    or_filter_active = len(selected_or_ids) < len(all_ors) if all_ors else False
    ward_filter_active = len(selected_ward_ids) < len(all_wards) if all_wards else False

    if or_filter_active or ward_filter_active:
        # Build OR filter (any of selected ORs)
        if or_filter_active:
            or_filter = Q(operating_room_id__in=selected_or_ids)
        else:
            or_filter = Q(operating_room__isnull=False)  # All ORs

        # Build Ward filter (any of selected Wards)
        if ward_filter_active:
            ward_filter = Q(ward_id__in=selected_ward_ids)
        else:
            ward_filter = Q(ward__isnull=False)  # All Wards

        # Combine: show messages from selected ORs AND selected Wards
        messages = messages.filter(or_filter & ward_filter)

    # Separate by message type
    can_accept_messages = messages.filter(message_type='CAN_ACCEPT_PATIENTS')
    patient_requested_messages = messages.filter(message_type='CAN_ACCEPT_PATIENTS')
    patient_in_or_messages = messages.filter(message_type='PATIENT_IN_THE_OR')
    surgery_done_messages = messages.filter(message_type='SURGERY_DONE')

    # Match OR busy periods (CAN_ACCEPT_PATIENTS -> PATIENT_IN_THE_OR -> SURGERY_DONE)
    journeys = match_or_busy_periods(patient_requested_messages, patient_in_or_messages, surgery_done_messages)

    # Sort ALL journeys by patient request time (sent_at) - ALWAYS
    journeys.sort(key=lambda j: j['patient_requested'].sent_at)

    # Group journeys by operating room
    # Note: Journeys are already sorted by time within each group
    journeys_by_or = {}
    for journey in journeys:
        or_id = journey['operating_room'].id
        or_name = journey['operating_room'].name

        if or_id not in journeys_by_or:
            journeys_by_or[or_id] = {
                'operating_room': journey['operating_room'],
                'journeys': []
            }

        journeys_by_or[or_id]['journeys'].append(journey)

    # Sort groups by Operating Room using sort field (with language fallback)
    def get_or_sort_key(item):
        or_obj = item[1]['operating_room']
        # First by sort field (None sorts last), then by translated name
        sort_value = or_obj.sort if hasattr(or_obj, 'sort') and or_obj.sort is not None else 999999
        name_value = getattr(or_obj, name_field, None) or or_obj.name
        return (sort_value, name_value)

    journeys_by_or = dict(sorted(
        journeys_by_or.items(),
        key=get_or_sort_key
    ))

    # Calculate statistics
    stats = calculate_statistics(journeys)

    # Get all dates with data for the current month (for calendar marking)
    # Use localtime to ensure we're working in the configured timezone
    from django.utils.timezone import localtime

    current_local_time = localtime(timezone.now())
    current_month_start = current_local_time.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if current_month_start.month == 12:
        next_month_start = current_month_start.replace(year=current_month_start.year + 1, month=1)
    else:
        next_month_start = current_month_start.replace(month=current_month_start.month + 1)

    # Get all dates with CAN_ACCEPT_PATIENTS messages in current month (respecting filters)
    # The .dates() method automatically converts to local timezone when USE_TZ=True
    dates_query = MessageLog.objects.filter(
        hospital=hospital,
        message_type='CAN_ACCEPT_PATIENTS',
        sent_at__gte=current_month_start,
        sent_at__lt=next_month_start
    )

    # Apply same filters to calendar dots
    if or_filter_active or ward_filter_active:
        if or_filter_active:
            or_filter_cal = Q(operating_room_id__in=selected_or_ids)
        else:
            or_filter_cal = Q(operating_room__isnull=False)

        if ward_filter_active:
            ward_filter_cal = Q(ward_id__in=selected_ward_ids)
        else:
            ward_filter_cal = Q(ward__isnull=False)

        dates_query = dates_query.filter(or_filter_cal & ward_filter_cal)

    dates_with_data = dates_query.dates('sent_at', 'day')

    # Convert to YYYY-MM-DD format for JavaScript
    dates_with_data_list = [d.strftime('%Y-%m-%d') for d in dates_with_data]

    # Calculate navigation dates
    if date_str:
        try:
            current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            current_date = timezone.now().date()
    else:
        current_date = timezone.now().date()

    # Calculate previous and next dates based on period
    if period == 'day':
        prev_date = current_date - timedelta(days=1)
        next_date = current_date + timedelta(days=1)
    elif period == 'week':
        prev_date = current_date - timedelta(weeks=1)
        next_date = current_date + timedelta(weeks=1)
    elif period == 'month':
        # Previous month
        if current_date.month == 1:
            prev_date = current_date.replace(year=current_date.year - 1, month=12)
        else:
            prev_date = current_date.replace(month=current_date.month - 1)
        # Next month
        if current_date.month == 12:
            next_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            next_date = current_date.replace(month=current_date.month + 1)
    elif period == 'quarter':
        prev_date = current_date - timedelta(days=90)
        next_date = current_date + timedelta(days=90)
    else:
        prev_date = current_date - timedelta(days=1)
        next_date = current_date + timedelta(days=1)

    # Determine if filters are active
    filters_active = or_filter_active or ward_filter_active

    context = {
        'hospital': hospital,
        'period': period,
        'date_str': date_str or current_date.strftime('%Y-%m-%d'),
        'display_name': display_name,
        'start_date': start_date,
        'end_date': end_date,
        'journeys_by_or': journeys_by_or,
        'stats': stats,
        'can_accept_count': can_accept_messages.count(),
        'patient_requested_count': patient_requested_messages.count(),
        'patient_in_or_count': patient_in_or_messages.count(),
        'surgery_done_count': surgery_done_messages.count(),
        'prev_date': prev_date.strftime('%Y-%m-%d'),
        'next_date': next_date.strftime('%Y-%m-%d'),
        'dates_with_data': mark_safe(json.dumps(dates_with_data_list)),
        # Filter-related context
        'all_ors': all_ors,
        'all_wards': all_wards,
        'selected_or_ids': selected_or_ids,
        'selected_ward_ids': selected_ward_ids,
        'filters_active': filters_active,
        'or_ids_param': ','.join(map(str, selected_or_ids)),
        'ward_ids_param': ','.join(map(str, selected_ward_ids)),
    }

    return render(request, 'stats/dashboard.html', context)


def export_xlsx(request, period='day', date_str=None):
    """
    Export statistics to XLSX file with filter support.
    """
    # Get current hospital from Site
    current_site = get_current_site(request)
    try:
        hospital = current_site.hospital
    except:
        return HttpResponse('No hospital configured for this site.', status=400)

    # Get date range
    start_date, end_date, display_name = get_date_range(period, date_str)

    # Get current language for sorting
    current_lang = get_language()

    # Determine which name field to use for sorting based on language
    if current_lang and current_lang.startswith('pl'):
        name_field = 'name_pl'
    else:
        name_field = 'name_en'

    # Get all operating rooms and wards for this hospital
    # Sort by 'sort' field first (nulls last), then by translated name
    from django.db.models import F
    all_ors = list(OperatingRoom.objects.filter(hospital=hospital).order_by(
        F('sort').asc(nulls_last=True), name_field, 'name'
    ))
    all_wards = list(Ward.objects.filter(hospital=hospital).order_by(
        F('sort').asc(nulls_last=True), name_field, 'name'
    ))

    # Parse filter parameters from URL (same as dashboard view)
    or_ids_param = request.GET.get('or_ids', '')
    ward_ids_param = request.GET.get('ward_ids', '')

    selected_or_ids = []
    if or_ids_param:
        try:
            selected_or_ids = [int(x) for x in or_ids_param.split(',') if x.strip().isdigit()]
        except (ValueError, AttributeError):
            selected_or_ids = []

    selected_ward_ids = []
    if ward_ids_param:
        try:
            selected_ward_ids = [int(x) for x in ward_ids_param.split(',') if x.strip().isdigit()]
        except (ValueError, AttributeError):
            selected_ward_ids = []

    # Default: if no filters specified, select all
    if not selected_or_ids:
        selected_or_ids = [or_obj.id for or_obj in all_ors]
    if not selected_ward_ids:
        selected_ward_ids = [ward.id for ward in all_wards]

    # Get all messages for this hospital in the date range
    messages = MessageLog.objects.filter(
        hospital=hospital,
        sent_at__gte=start_date,
        sent_at__lt=end_date
    ).select_related('sender_role', 'recipient_role')

    # Apply filters (same logic as dashboard)
    or_filter_active = len(selected_or_ids) < len(all_ors) if all_ors else False
    ward_filter_active = len(selected_ward_ids) < len(all_wards) if all_wards else False

    if or_filter_active or ward_filter_active:
        if or_filter_active:
            or_filter = Q(operating_room_id__in=selected_or_ids)
        else:
            or_filter = Q(operating_room__isnull=False)

        if ward_filter_active:
            ward_filter = Q(ward_id__in=selected_ward_ids)
        else:
            ward_filter = Q(ward__isnull=False)

        messages = messages.filter(or_filter & ward_filter)

    # Separate by message type
    can_accept_messages = messages.filter(message_type='CAN_ACCEPT_PATIENTS')
    patient_requested_messages = messages.filter(message_type='CAN_ACCEPT_PATIENTS')
    patient_in_or_messages = messages.filter(message_type='PATIENT_IN_THE_OR')
    surgery_done_messages = messages.filter(message_type='SURGERY_DONE')

    # Match OR busy periods (CAN_ACCEPT_PATIENTS -> PATIENT_IN_THE_OR -> SURGERY_DONE)
    journeys = match_or_busy_periods(patient_requested_messages, patient_in_or_messages, surgery_done_messages)

    # Sort ALL journeys by patient request time (sent_at) - ALWAYS
    journeys.sort(key=lambda j: j['patient_requested'].sent_at)

    # Group journeys by operating room
    journeys_by_or = {}
    for journey in journeys:
        or_id = journey['operating_room'].id
        or_name = journey['operating_room'].name

        if or_id not in journeys_by_or:
            journeys_by_or[or_id] = {
                'operating_room': journey['operating_room'],
                'journeys': []
            }

        journeys_by_or[or_id]['journeys'].append(journey)

    # Sort groups by Operating Room using sort field (with language fallback)
    def get_or_sort_key(item):
        or_obj = item[1]['operating_room']
        # First by sort field (None sorts last), then by translated name
        sort_value = or_obj.sort if hasattr(or_obj, 'sort') and or_obj.sort is not None else 999999
        name_value = getattr(or_obj, name_field, None) or or_obj.name
        return (sort_value, name_value)

    journeys_by_or = dict(sorted(
        journeys_by_or.items(),
        key=get_or_sort_key
    ))

    # Calculate statistics
    stats = calculate_statistics(journeys)

    # Create Excel workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    summary_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # =================================================================
    # SHEET 1: Message Log (FIRST SHEET - user requirement)
    # =================================================================
    ws_log = wb.create_sheet(_("Message Log"), 0)  # Index 0 = first sheet

    # Header with filter info if applicable
    ws_log['A1'] = _("Complete Message Log")
    ws_log['A1'].font = Font(bold=True, size=14)

    filter_info = f"{display_name} - {hospital.name}"
    if or_filter_active or ward_filter_active:
        filter_info += " - " + _("Filtered")
    ws_log['A2'] = filter_info

    # Get all messages ordered by time
    all_messages = messages.order_by('sent_at').select_related(
        'sender_role',
        'recipient_role'
    )

    # Table headers (NEW: Added Operating Room and Ward columns)
    log_headers = [
        _("Timestamp"),
        _("Message Type"),
        _("Sender Role"),
        _("Recipient Role"),
        _("Operating Room"),  # NEW COLUMN
        _("Ward"),             # NEW COLUMN
        _("Acknowledged At"),
        _("Duration to Ack (sec)")
    ]

    header_row = 4
    for col, header in enumerate(log_headers, start=1):
        cell = ws_log.cell(row=header_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    # Data rows
    for row_idx, msg in enumerate(all_messages, start=header_row + 1):
        # Timestamp
        ws_log.cell(row=row_idx, column=1, value=msg.sent_at.strftime('%Y-%m-%d %H:%M:%S'))

        # Message Type
        msg_type_display = msg.get_message_type_display() if hasattr(msg, 'get_message_type_display') else msg.message_type
        ws_log.cell(row=row_idx, column=2, value=msg_type_display)

        # Sender Role
        sender_role = msg.sender_role.name if msg.sender_role else "N/A"
        ws_log.cell(row=row_idx, column=3, value=sender_role)

        # Recipient Role
        recipient_role = msg.recipient_role.name if msg.recipient_role else "N/A"
        ws_log.cell(row=row_idx, column=4, value=recipient_role)

        # Operating Room / Ward columns (NEW)
        or_name = ""
        ward_name = ""
        try:
            operating_room = OperatingRoom.objects.get(id=msg.operating_room_id)
            or_name = get_translated_name(operating_room)
        except OperatingRoom.DoesNotExist:
            or_name = f"ID: {msg.operating_room_id}"

        try:
            ward = Ward.objects.get(id=msg.ward_id)
            ward_name = get_translated_name(ward)
        except Ward.DoesNotExist:
            ward_name = f"ID: {msg.ward_id}"

        ws_log.cell(row=row_idx, column=5, value=or_name)
        ws_log.cell(row=row_idx, column=6, value=ward_name)

        # Acknowledged At
        if msg.acknowledged_at:
            ws_log.cell(row=row_idx, column=7, value=msg.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S'))
            # Duration to acknowledge (in seconds)
            duration = (msg.acknowledged_at - msg.sent_at).total_seconds()
            ws_log.cell(row=row_idx, column=8, value=int(duration))
        else:
            ws_log.cell(row=row_idx, column=7, value=_("Not acknowledged"))
            ws_log.cell(row=row_idx, column=8, value="N/A")

        # Apply borders to all cells
        for col in range(1, 9):
            ws_log.cell(row=row_idx, column=col).border = border

    # Add Excel Table with auto-filters
    if len(all_messages) > 0:
        from openpyxl.worksheet.table import Table, TableStyleInfo
        last_row = header_row + len(all_messages)
        table_ref = f"A{header_row}:H{last_row}"
        table = Table(displayName="MessageLog", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws_log.add_table(table)

    # Auto-size columns
    ws_log.column_dimensions['A'].width = 20
    ws_log.column_dimensions['B'].width = 20
    ws_log.column_dimensions['C'].width = 15
    ws_log.column_dimensions['D'].width = 15
    ws_log.column_dimensions['E'].width = 25  # Operating Room
    ws_log.column_dimensions['F'].width = 25  # Ward
    ws_log.column_dimensions['G'].width = 20
    ws_log.column_dimensions['H'].width = 20

    # =================================================================
    # SHEET 2: Summary sheet
    # =================================================================
    ws_summary = wb.create_sheet(_("Summary"))
    ws_summary['A1'] = _("Operating Room Statistics")
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A2'] = hospital.name
    ws_summary['A3'] = display_name

    # Summary statistics
    summary_row = 5
    ws_summary[f'A{summary_row}'] = _("Statistic")
    ws_summary[f'B{summary_row}'] = _("Value")
    for cell in [ws_summary[f'A{summary_row}'], ws_summary[f'B{summary_row}']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    summary_data = [
        (_("Completed Surgeries"), stats['total_completed']),
        (_("Incomplete Entries"), stats['total_incomplete']),
        (_("Total OR Time Used (hours)"), stats['total_hours']),
        (_("Average OR Time (minutes)"), stats['avg_minutes']),
        (_("Patients Requested"), can_accept_messages.count()),
    ]

    for idx, (label, value) in enumerate(summary_data, start=summary_row + 1):
        ws_summary[f'A{idx}'] = label
        ws_summary[f'B{idx}'] = value
        for cell in [ws_summary[f'A{idx}'], ws_summary[f'B{idx}']]:
            cell.border = border
        ws_summary[f'A{idx}'].fill = summary_fill

    # Auto-size columns
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15

    # Create detailed sheets for each operating room
    for or_id, or_data in journeys_by_or.items():
        or_name = get_translated_name(or_data['operating_room'])

        # Sanitize sheet name - Excel doesn't allow: \ / ? * [ ] :
        safe_sheet_name = or_name
        invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            safe_sheet_name = safe_sheet_name.replace(char, '_')

        # Sheet names limited to 31 chars
        safe_sheet_name = safe_sheet_name[:31]

        ws = wb.create_sheet(safe_sheet_name)

        # Header
        ws['A1'] = or_name
        ws['A1'].font = Font(bold=True, size=14)

        # Table headers
        headers = [
            _("Status"),
            _("Patient Requested"),
            _("Patient in OR"),
            _("Surgery Completed"),
            _("Time from last patient (min)"),
            _("Wait for Patient (min)"),
            _("Operation Time (min)"),
            _("Total OR Time (min)"),
            _("Ward")
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # Data rows
        for row_idx, journey in enumerate(or_data['journeys'], start=4):
            # Status
            status = _("Complete") if not journey['incomplete'] else _("Incomplete")
            ws.cell(row=row_idx, column=1, value=status)

            # Patient Requested
            requested_time = journey['patient_requested'].sent_at.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_idx, column=2, value=requested_time)

            # Patient in OR
            if journey['patient_in_or']:
                patient_in_or_time = journey['patient_in_or'].sent_at.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row_idx, column=3, value=patient_in_or_time)
            else:
                ws.cell(row=row_idx, column=3, value="N/A")

            # Surgery Completed
            if journey['surgery_done']:
                completed_time = journey['surgery_done'].sent_at.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row_idx, column=4, value=completed_time)
            else:
                ws.cell(row=row_idx, column=4, value=_("Not completed"))

            # Time Between Calls
            if journey['time_between_calls_minutes'] is not None:
                ws.cell(row=row_idx, column=5, value=journey['time_between_calls_minutes'])
            else:
                ws.cell(row=row_idx, column=5, value="N/A")

            # Wait for Patient
            if journey['wait_time_minutes'] is not None:
                ws.cell(row=row_idx, column=6, value=journey['wait_time_minutes'])
            else:
                ws.cell(row=row_idx, column=6, value="N/A")

            # Operation Time
            if journey['operation_time_minutes'] is not None:
                ws.cell(row=row_idx, column=7, value=journey['operation_time_minutes'])
            else:
                ws.cell(row=row_idx, column=7, value="N/A")

            # Total OR Time
            if journey['total_time_minutes'] is not None:
                ws.cell(row=row_idx, column=8, value=journey['total_time_minutes'])
            else:
                ws.cell(row=row_idx, column=8, value="N/A")

            # Ward
            ward_name = "N/A"
            try:
                ward = Ward.objects.get(id=journey['patient_requested'].ward_id)
                ward_name = get_translated_name(ward)
            except Ward.DoesNotExist:
                ward_name = f"Ward ID: {journey['patient_requested'].ward_id}"
            ws.cell(row=row_idx, column=9, value=ward_name)

            # Apply borders
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).border = border

        # Add Excel Table with auto-filters for this OR sheet
        if len(or_data['journeys']) > 0:
            from openpyxl.worksheet.table import Table, TableStyleInfo
            last_row = 3 + len(or_data['journeys'])
            table_ref = f"A3:I{last_row}"
            # Use unique table name for each OR (sanitize OR name for table name)
            table_name = safe_sheet_name.replace(' ', '').replace('-', '').replace('_', '')[:31]
            # Ensure table name starts with a letter
            if not table_name[0].isalpha():
                table_name = 'OR' + table_name
            table = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        # Auto-size columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20  # Time Between Calls
        ws.column_dimensions['F'].width = 20  # Wait for Patient
        ws.column_dimensions['G'].width = 20  # Operation Time
        ws.column_dimensions['H'].width = 20  # Total OR Time
        ws.column_dimensions['I'].width = 20  # Ward

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Create nice filename with format: report_<period>_<date-range>_<hospital>.xlsx
    # Examples:
    #   report_day_2024-01-15_Hospital.xlsx
    #   report_week_2024-01-15_to_2024-01-21_Hospital.xlsx
    #   report_month_January_2024_Hospital.xlsx
    #   report_quarter_Q1_2024_Hospital.xlsx

    # Sanitize display name for filename - remove all special chars
    import re
    safe_display_name = re.sub(r'[^\w\-]', '_', display_name)
    safe_hospital_name = re.sub(r'[^\w\-]', '_', hospital.name)

    filename = f"report_{period}_{safe_display_name}_{safe_hospital_name}.xlsx"

    # Use both filename and filename* (RFC 6266) for better browser compatibility
    from urllib.parse import quote
    encoded_filename = quote(filename)
    response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'

    wb.save(response)
    return response
